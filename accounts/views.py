from io import BytesIO

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils.crypto import get_random_string
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from academics.models import Enrollment, Subject
from portfolios.models import Portfolio

from .forms import BulkStudentAccountForm
from .models import StudentProfile, User


SECTION_CODES = {
    "أ": "a",
    "ا": "a",
    "ب": "b",
    "ج": "c",
    "د": "d",
    "هـ": "e",
}


@require_GET
def home(request):
    """عرض الصفحة الرئيسية لجميع المستخدمين."""

    return render(
        request,
        "accounts/home.html",
    )


@require_GET
def parent_inquiry(request):
    """الانتقال إلى صفحة استعلام ولي الأمر الخارجية."""

    inquiry_url = settings.PARENT_INQUIRY_URL

    if not inquiry_url.startswith("https://"):
        return redirect("accounts:home")

    return redirect(inquiry_url)


@login_required
@require_GET
def dashboard_redirect(request):
    """توجيه المستخدم إلى لوحته حسب نوع حسابه."""

    user = request.user

    if user.is_superuser:
        return redirect("accounts:teacher_dashboard")

    if user.role == User.Role.ADMIN:
        return redirect("admin:index")

    if user.role == User.Role.TEACHER:
        return redirect("accounts:teacher_dashboard")

    if user.role == User.Role.STUDENT:
        return redirect("accounts:student_dashboard")

    if user.role == User.Role.PARENT:
        return redirect("accounts:parent_dashboard")

    return redirect("accounts:home")


@login_required
@require_GET
def teacher_dashboard(request):
    """عرض لوحة المعلمة."""

    if (
        request.user.role != User.Role.TEACHER
        and not request.user.is_superuser
    ):
        raise PermissionDenied(
            "ليس لديك صلاحية لدخول لوحة المعلمة."
        )

    context = {
        "page_title": "لوحة المعلمة",
        "user": request.user,
    }

    return render(
        request,
        "accounts/teacher_dashboard.html",
        context,
    )


@login_required
@require_GET
def student_dashboard(request):
    """عرض لوحة الطالبة."""

    if (
        request.user.role != User.Role.STUDENT
        and not request.user.is_superuser
    ):
        raise PermissionDenied(
            "ليس لديك صلاحية لدخول لوحة الطالبة."
        )

    context = {
        "page_title": "دفتر العلوم الخاص بي",
        "user": request.user,
    }

    return render(
        request,
        "accounts/student_dashboard.html",
        context,
    )


@login_required
@require_GET
def parent_dashboard(request):
    """عرض لوحة ولي الأمر."""

    if (
        request.user.role != User.Role.PARENT
        and not request.user.is_superuser
    ):
        raise PermissionDenied(
            "ليس لديك صلاحية لدخول لوحة ولي الأمر."
        )

    context = {
        "page_title": "متابعة الطالبة",
        "user": request.user,
    }

    return render(
        request,
        "accounts/parent_dashboard.html",
        context,
    )


def _student_username(classroom, index):
    """إنشاء اسم مستخدم آمن حسب الصف والشعبة."""

    grade_code = classroom.grade_level.code.lower()
    section_code = SECTION_CODES.get(
        classroom.name.strip(),
        "x",
    )

    return f"sci_{grade_code}{section_code}_{index:03d}"


def _temporary_password():
    """إنشاء كلمة مرور مؤقتة قوية."""

    allowed_characters = (
        "abcdefghjkmnpqrstuvwxyz"
        "ABCDEFGHJKLMNPQRSTUVWXYZ"
        "23456789"
        "!@#$"
    )

    return get_random_string(
        12,
        allowed_characters,
    )


def _create_credentials_workbook(classroom, credentials):
    """إنشاء كشف Excel ببيانات الحسابات."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "حسابات الطالبات"
    worksheet.sheet_view.rightToLeft = True

    headers = [
        "م",
        "الصف والشعبة",
        "اسم الطالبة",
        "اسم المستخدم",
        "كلمة المرور المؤقتة",
        "حالة الحساب",
    ]

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0F766E",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for index, credential in enumerate(
        credentials,
        start=1,
    ):
        worksheet.append(
            [
                index,
                str(classroom),
                credential["student_name"],
                credential["username"],
                credential["password"],
                credential["status"],
            ]
        )

    column_widths = {
        "A": 8,
        "B": 25,
        "C": 25,
        "D": 22,
        "E": 25,
        "F": 20,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


@login_required
@require_http_methods(["GET", "POST"])
def bulk_create_students(request):
    """إنشاء حسابات الطالبات وربطها بالشعبة والدفاتر."""

    if (
        request.user.role != User.Role.TEACHER
        and not request.user.is_superuser
    ):
        raise PermissionDenied(
            "ليس لديك صلاحية لإنشاء حسابات الطالبات."
        )

    form = BulkStudentAccountForm(
        request.POST or None,
        user=request.user,
    )

    if request.method == "POST" and form.is_valid():
        classroom = form.cleaned_data["classroom"]
        student_count = form.cleaned_data["student_count"]

        science_subject = Subject.objects.filter(
            code="SCI",
            is_active=True,
        ).first()

        if science_subject is None:
            form.add_error(
                None,
                "لم يتم العثور على مقرر العلوم النشط.",
            )
        else:
            credentials = []

            with transaction.atomic():
                for index in range(1, student_count + 1):
                    username = _student_username(
                        classroom,
                        index,
                    )

                    user = User.objects.filter(
                        username=username,
                    ).first()

                    if user is None:
                        password = _temporary_password()

                        user = User(
                            username=username,
                            first_name="طالبة",
                            last_name=f"{index:03d}",
                            role=User.Role.STUDENT,
                            email=None,
                            is_active=True,
                            is_verified=False,
                        )

                        user.set_password(password)
                        user.save()

                        account_status = "تم إنشاء الحساب"
                    else:
                        password = ""
                        account_status = "موجود مسبقًا"

                        if user.role != User.Role.STUDENT:
                            raise PermissionDenied(
                                f"اسم المستخدم {username} "
                                "مرتبط بحساب غير مخصص لطالبة."
                            )

                    StudentProfile.objects.get_or_create(
                        user=user,
                        defaults={
                            "student_number": username,
                        },
                    )

                    Enrollment.objects.get_or_create(
                        student=user,
                        classroom=classroom,
                        defaults={
                            "status": Enrollment.Status.ACTIVE,
                        },
                    )

                    Portfolio.objects.get_or_create(
                        student=user,
                        classroom=classroom,
                        subject=science_subject,
                        academic_year=classroom.academic_year,
                        defaults={
                            "title": "دفتر العلوم الذكي",
                            "is_active": True,
                        },
                    )

                    credentials.append(
                        {
                            "student_name": (
                                f"طالبة {index:03d}"
                            ),
                            "username": username,
                            "password": password,
                            "status": account_status,
                        }
                    )

            excel_file = _create_credentials_workbook(
                classroom,
                credentials,
            )

            filename = (
                f"student_accounts_"
                f"{classroom.grade_level.code}_"
                f"{classroom.id}.xlsx"
            )

            response = HttpResponse(
                excel_file.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
            )

            response["Content-Disposition"] = (
                f'attachment; filename="{filename}"'
            )

            response["X-Content-Type-Options"] = "nosniff"

            return response

    context = {
        "page_title": "إنشاء حسابات الطالبات",
        "form": form,
    }

    return render(
        request,
        "accounts/bulk_create_students.html",
        context,
    )