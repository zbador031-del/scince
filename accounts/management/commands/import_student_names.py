import json
from collections import defaultdict
from pathlib import Path

from django.conf import settings
from django.core.management.base import (
    BaseCommand,
    CommandError,
)
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from academics.models import (
    AcademicYear,
    Classroom,
    Enrollment,
)


REQUIRED_COLUMNS = {
    "اسم الطالبة الكامل",
    "الصف",
    "الفصل",
}


class Command(BaseCommand):
    """
    تحديث أسماء حسابات الطالبات من ملف Excel.

    يعمل افتراضيًا بوضع المعاينة ولا يغيّر قاعدة البيانات
    إلا عند استخدام الخيار --apply.
    """

    help = (
        "مطابقة أسماء الطالبات مع الحسابات الحالية، "
        "وتحديث الأسماء دون تغيير بيانات الدخول."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="student_names_import.xlsx",
            help="مسار ملف Excel.",
        )

        parser.add_argument(
            "--apply",
            action="store_true",
            help="تطبيق التحديث فعليًا.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])

        if not file_path.is_absolute():
            file_path = (
                Path(settings.BASE_DIR) / file_path
            )

        if not file_path.exists():
            raise CommandError(
                f"لم يتم العثور على الملف: {file_path}"
            )

        academic_year = self.get_current_academic_year()

        roster_groups = self.read_roster(file_path)

        mappings = self.build_mappings(
            roster_groups=roster_groups,
            academic_year=academic_year,
        )

        self.show_preview(
            mappings=mappings,
            academic_year=academic_year,
        )

        if not options["apply"]:
            self.stdout.write("")

            self.stdout.write(
                self.style.WARNING(
                    "هذه معاينة فقط، ولم يتم تعديل "
                    "قاعدة البيانات."
                )
            )

            self.stdout.write(
                "بعد مراجعة المطابقة نفّذي الأمر "
                "مرة أخرى مع إضافة --apply."
            )

            return

        backup_path = self.create_backup(mappings)

        self.apply_updates(mappings)

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                f"تم تحديث أسماء {len(mappings)} "
                f"طالبة بنجاح."
            )
        )

        self.stdout.write(
            self.style.SUCCESS(
                f"تم حفظ النسخة الاحتياطية في: "
                f"{backup_path}"
            )
        )

        self.stdout.write(
            self.style.WARNING(
                "الحسابات الزائدة تُركت دون تعديل "
                "للمراجعة اليدوية."
            )
        )

    def get_current_academic_year(self):
        """الحصول على العام الدراسي الحالي."""

        try:
            return AcademicYear.objects.get(
                is_current=True
            )

        except AcademicYear.DoesNotExist as error:
            raise CommandError(
                "لم يتم تحديد عام دراسي حالي "
                "في لوحة الإدارة."
            ) from error

        except AcademicYear.MultipleObjectsReturned as error:
            raise CommandError(
                "يوجد أكثر من عام دراسي محدد "
                "بوصفه عامًا حاليًا."
            ) from error

    def read_roster(self, file_path):
        """قراءة ملف Excel والتحقق من البيانات."""

        try:
            workbook = load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True,
            )

        except Exception as error:
            raise CommandError(
                f"تعذر فتح ملف Excel: {error}"
            ) from error

        sheet_name = "قائمة الاستيراد"

        if sheet_name not in workbook.sheetnames:
            workbook.close()

            raise CommandError(
                f"لا توجد ورقة باسم: {sheet_name}"
            )

        worksheet = workbook[sheet_name]

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

        column_positions = {
            self.clean_text(value): index
            for index, value in enumerate(header_row)
            if value is not None
        }

        missing_columns = (
            REQUIRED_COLUMNS - set(column_positions)
        )

        if missing_columns:
            workbook.close()

            missing_names = "، ".join(
                sorted(missing_columns)
            )

            raise CommandError(
                f"الأعمدة التالية غير موجودة: "
                f"{missing_names}"
            )

        roster_groups = defaultdict(list)
        duplicate_check = set()

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            full_name = self.clean_text(
                row[
                    column_positions[
                        "اسم الطالبة الكامل"
                    ]
                ]
            )

            grade_name = self.clean_text(
                row[column_positions["الصف"]]
            )

            section_name = self.clean_section(
                row[column_positions["الفصل"]]
            )

            if not full_name:
                continue

            if not grade_name or not section_name:
                workbook.close()

                raise CommandError(
                    f"بيانات الصف أو الفصل ناقصة "
                    f"في السطر {row_number}."
                )

            duplicate_key = (
                grade_name,
                section_name,
                full_name,
            )

            if duplicate_key in duplicate_check:
                workbook.close()

                raise CommandError(
                    f"اسم مكرر في السطر "
                    f"{row_number}: {full_name}"
                )

            duplicate_check.add(duplicate_key)

            roster_groups[
                (grade_name, section_name)
            ].append(
                {
                    "full_name": full_name,
                    "row_number": row_number,
                }
            )

        workbook.close()

        if not roster_groups:
            raise CommandError(
                "لم يتم العثور على أسماء طالبات."
            )

        return roster_groups

    def build_mappings(
        self,
        roster_groups,
        academic_year,
    ):
        """مطابقة الأسماء مع حسابات كل فصل."""

        mappings = []

        for (
            grade_name,
            section_name,
        ), roster_students in roster_groups.items():

            grade_keyword = grade_name.replace(
                "الابتدائي",
                "",
            ).strip()

            classrooms = (
                Classroom.objects.filter(
                    academic_year=academic_year,
                    grade_level__name__icontains=(
                        grade_keyword
                    ),
                    name=section_name,
                    is_active=True,
                )
                .select_related(
                    "grade_level",
                    "academic_year",
                )
            )

            classroom_count = classrooms.count()

            if classroom_count == 0:
                raise CommandError(
                    f"لم يتم العثور على الفصل: "
                    f"{grade_name} / {section_name}"
                )

            if classroom_count > 1:
                raise CommandError(
                    f"يوجد أكثر من فصل مطابق لـ: "
                    f"{grade_name} / {section_name}"
                )

            classroom = classrooms.first()

            enrollments = list(
                Enrollment.objects.filter(
                    classroom=classroom,
                    status=Enrollment.Status.ACTIVE,
                    student__role="student",
                )
                .select_related("student")
                .order_by(
                    "student__username",
                    "pk",
                )
            )

            roster_count = len(roster_students)
            account_count = len(enrollments)

            if roster_count > account_count:
                raise CommandError(
                    f"عدد الأسماء في {classroom} "
                    f"هو {roster_count}، بينما عدد "
                    f"الحسابات هو {account_count}."
                )

            extra_accounts = (
                account_count - roster_count
            )

            self.stdout.write("")

            self.stdout.write(
                self.style.MIGRATE_HEADING(
                    f"{classroom}: "
                    f"{roster_count} اسمًا، "
                    f"{account_count} حسابًا، "
                    f"{extra_accounts} حساب زائد."
                )
            )

            for roster_student, enrollment in zip(
                roster_students,
                enrollments,
            ):
                first_name, last_name = (
                    self.split_full_name(
                        roster_student["full_name"]
                    )
                )

                mappings.append(
                    {
                        "student": enrollment.student,
                        "classroom": classroom,
                        "first_name": first_name,
                        "last_name": last_name,
                        "full_name": (
                            roster_student["full_name"]
                        ),
                        "excel_row": (
                            roster_student["row_number"]
                        ),
                    }
                )

        return mappings

    def show_preview(
        self,
        mappings,
        academic_year,
    ):
        """عرض المطابقة قبل التحديث."""

        self.stdout.write("")

        self.stdout.write(
            self.style.SUCCESS(
                f"معاينة المطابقة للعام: "
                f"{academic_year}"
            )
        )

        current_classroom = None

        for mapping in mappings:
            classroom = mapping["classroom"]

            if classroom != current_classroom:
                current_classroom = classroom

                self.stdout.write("")

                self.stdout.write(
                    self.style.MIGRATE_LABEL(
                        f"الفصل: {classroom}"
                    )
                )

            student = mapping["student"]

            self.stdout.write(
                f"{student.username}  ←  "
                f"{mapping['full_name']}"
            )

        self.stdout.write("")

        self.stdout.write(
            f"إجمالي الحسابات المطابقة: "
            f"{len(mappings)}"
        )

    def create_backup(self, mappings):
        """إنشاء نسخة احتياطية من الأسماء الحالية."""

        backup_directory = (
            Path(settings.BASE_DIR) / "backups"
        )

        backup_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        timestamp = timezone.localtime().strftime(
            "%Y%m%d_%H%M%S"
        )

        backup_path = (
            backup_directory
            / f"student_names_before_{timestamp}.json"
        )

        backup_data = {
            "created_at": timezone.now().isoformat(),
            "student_count": len(mappings),
            "students": [
                {
                    "id": mapping["student"].pk,
                    "username": (
                        mapping["student"].username
                    ),
                    "first_name": (
                        mapping["student"].first_name
                    ),
                    "last_name": (
                        mapping["student"].last_name
                    ),
                    "classroom": str(
                        mapping["classroom"]
                    ),
                }
                for mapping in mappings
            ],
        }

        backup_path.write_text(
            json.dumps(
                backup_data,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        return backup_path

    @transaction.atomic
    def apply_updates(self, mappings):
        """تحديث الأسماء مع إبقاء بيانات الدخول."""

        for mapping in mappings:
            student = mapping["student"]

            student.first_name = (
                mapping["first_name"]
            )

            student.last_name = (
                mapping["last_name"]
            )

            student.save(
                update_fields=[
                    "first_name",
                    "last_name",
                ]
            )

    @staticmethod
    def clean_text(value):
        """تنظيف المسافات الزائدة."""

        if value is None:
            return ""

        return " ".join(
            str(value).strip().split()
        )

    @staticmethod
    def clean_section(value):
        """تحويل رقم الفصل إلى نص."""

        if value is None:
            return ""

        if (
            isinstance(value, float)
            and value.is_integer()
        ):
            return str(int(value))

        return str(value).strip()

    @staticmethod
    def split_full_name(full_name):
        """تقسيم الاسم الكامل."""

        parts = full_name.split(maxsplit=1)

        first_name = parts[0]

        last_name = (
            parts[1]
            if len(parts) > 1
            else ""
        )

        return first_name, last_name